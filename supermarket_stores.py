import requests
from bs4 import BeautifulSoup
from pathlib import Path
import openpyxl

# excelブック/シートの準備
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Stores'
ws.cell(row=1, column=1,value="地域")
ws.cell(row=1, column=2,value="市区町村")
ws.cell(row=1, column=3,value="スーパー名")
ws.cell(row=1, column=4,value="店名")

cnt = 2
for i in range(47):                                                                     # 47都道府県を検索
    url = f'https://www.homemate-research-supermarket.com/search-list/{str(i+1).zfill(2)}/'
    soup = BeautifulSoup(requests.get(url).content, 'html.parser')
    areas = soup.select('section.inner')

    for area in areas:
        cities = area.select('[class="ttl3 ttl3-2"]')
        area_stores = area.select('ul.areaul')
        for num, city_stores in enumerate(area_stores):
            city_stores = city_stores.select('a')
            for city_store in city_stores:
                ws.cell(row=cnt, column=1, value=area.h2.span.text.split('（')[0])      # 地域
                ws.cell(row=cnt, column=2, value=cities[num].text.split('（')[0])       # 市区町村
                store = city_store.text.split('\u3000')
                ws.cell(row=cnt, column=3, value=store[0])                              # スーパー名
                if len(store) > 1:
                    ws.cell(row=cnt, column=4, value=store[1])                          # 店鋪名
                else:
                    ws.cell(row=cnt, column=4, value="*")
                cnt += 1

# フォルダーを作成
folder = Path('output')
folder.mkdir(exist_ok=True)

# excelファイルに出力
excel_path = folder.joinpath('SupermarketStores.xlsx')
wb.save(excel_path)
wb.close()
